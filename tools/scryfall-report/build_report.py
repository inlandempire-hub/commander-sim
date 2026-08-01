"""
Reads ./data/oracle-cards.jsonl.gz (see fetch_bulk_data.py), filters to
Commander-legal cards, tags each with a heuristic archetype guess and
complexity score, and writes ./output/commander_card_report.xlsx.
"""

import gzip
import json
import os
import re

import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from heuristics import MULTI_FACED_LAYOUTS, score_complexity, tag_archetypes

HERE = os.path.dirname(__file__)
DATA_PATH = os.path.join(HERE, "data", "oracle-cards.jsonl.gz")
OUTPUT_DIR = os.path.join(HERE, "output")
OUTPUT_PATH = os.path.join(OUTPUT_DIR, "commander_card_report.xlsx")
ENGINE_TEST_CARDS_PATH = os.path.join(
    HERE, "..", "..", "packages", "engine", "src", "cards", "testCards.ts"
)


def load_already_implemented_names() -> set[str]:
    """Parses the engine's test card fixtures for real (non-placeholder) card names."""
    with open(ENGINE_TEST_CARDS_PATH, "r", encoding="utf-8") as f:
        source = f.read()

    names = set()
    for match in re.finditer(r'id:\s*"([^"]+)",\s*\n\s*name:\s*"([^"]+)"', source):
        card_id, name = match.group(1), match.group(2)
        if card_id.startswith("test-commander"):
            continue  # placeholder fixtures, not real cards
        names.add(name.lower())
    return names


def combined_oracle_text(card: dict) -> str:
    if card.get("oracle_text"):
        return card["oracle_text"]
    faces = card.get("card_faces") or []
    return " // ".join(face.get("oracle_text", "") for face in faces if face.get("oracle_text"))


def combined_mana_cost(card: dict) -> str:
    if card.get("mana_cost"):
        return card["mana_cost"]
    faces = card.get("card_faces") or []
    costs = [face.get("mana_cost", "") for face in faces if face.get("mana_cost")]
    return " // ".join(costs)


def combined_power_toughness(card: dict) -> tuple[str, str]:
    if "power" in card or "toughness" in card:
        return card.get("power", ""), card.get("toughness", "")
    faces = card.get("card_faces") or []
    if faces and "power" in faces[0]:
        return faces[0].get("power", ""), faces[0].get("toughness", "")
    return "", ""


def extract_row(card: dict) -> dict:
    layout = card.get("layout", "")
    is_multi_faced = layout in MULTI_FACED_LAYOUTS
    oracle_text = combined_oracle_text(card)
    power, toughness = combined_power_toughness(card)
    keywords = card.get("keywords") or []

    complexity = score_complexity(
        oracle_text=oracle_text,
        type_line=card.get("type_line", ""),
        is_multi_faced=is_multi_faced,
        keywords=keywords,
    )
    archetypes = tag_archetypes(oracle_text)

    return {
        "name": card.get("name", ""),
        "type_line": card.get("type_line", ""),
        "mana_cost": combined_mana_cost(card),
        "cmc": card.get("cmc", 0),
        "colors": "".join(card.get("colors") or []),
        "color_identity": "".join(card.get("color_identity") or []),
        "power": power,
        "toughness": toughness,
        "keywords": ", ".join(keywords),
        "rarity": card.get("rarity", ""),
        "oracle_text": oracle_text,
        "archetype_tags": "; ".join(archetypes) if archetypes else "",
        "complexity_score": complexity.score,
        "complexity_tier": complexity.tier,
        "complexity_flags": "; ".join(complexity.flags),
        "is_multi_faced": is_multi_faced,
        "layout": layout,
    }


def load_commander_legal_rows() -> list[dict]:
    rows = []
    with gzip.open(DATA_PATH, "rt", encoding="utf-8") as f:
        for line in f:
            card = json.loads(line)
            if card.get("legalities", {}).get("commander") == "legal":
                rows.append(extract_row(card))
    return rows


def write_xlsx(df: pd.DataFrame) -> None:
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    df.to_excel(OUTPUT_PATH, sheet_name="Commander card pool", index=False, engine="openpyxl")

    from openpyxl import load_workbook

    wb = load_workbook(OUTPUT_PATH)
    ws = wb["Commander card pool"]

    header_font = Font(name="Arial", bold=True)
    body_font = Font(name="Arial")

    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=False, vertical="center")

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    column_widths = {
        "name": 28,
        "type_line": 26,
        "mana_cost": 12,
        "cmc": 6,
        "colors": 8,
        "color_identity": 10,
        "power": 6,
        "toughness": 8,
        "keywords": 22,
        "rarity": 10,
        "oracle_text": 60,
        "archetype_tags": 32,
        "complexity_score": 10,
        "complexity_tier": 14,
        "complexity_flags": 45,
        "is_multi_faced": 10,
        "layout": 12,
        "already_implemented": 12,
    }
    for idx, column_name in enumerate(df.columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = column_widths.get(column_name, 15)

    wb.save(OUTPUT_PATH)


def main() -> None:
    if not os.path.exists(DATA_PATH):
        raise SystemExit(f"Missing {DATA_PATH} - run fetch_bulk_data.py first")

    already_implemented = load_already_implemented_names()
    rows = load_commander_legal_rows()
    df = pd.DataFrame(rows)
    df["already_implemented"] = df["name"].str.lower().isin(already_implemented)

    df = df.sort_values(["complexity_score", "name"]).reset_index(drop=True)

    write_xlsx(df)

    print(f"{len(df):,} Commander-legal cards written to {OUTPUT_PATH}")
    print(df["complexity_tier"].value_counts().sort_index())
    print(f"Already implemented: {int(df['already_implemented'].sum())}")


if __name__ == "__main__":
    main()
